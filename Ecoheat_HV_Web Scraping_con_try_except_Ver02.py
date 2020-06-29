import random
import xlsxwriter
import time
from time import sleep
from openpyxl import load_workbook
from selenium import webdriver
from tkinter import filedialog
from tkinter import *

# Lista de Variables
root = Tk()
Final_dict = {}
Salida_keys = []
Salida_Values = []
numero_credito = []
monto_vale = []
lista_de_creditos = []
contador = 0

# Cargamos el controlador y definimos su ruta de acceso
chromedriver_location = '/usr/local/bin/chromedriver'
driver = webdriver.Chrome(chromedriver_location)
# abrimos la pagina
driver.get('')
# esperamos segundo y medio
sleep(1.5)

# login
# Es necesario crear un login

# Definimos los campos a usar
username_input = '//*[@id="login_wrapper"]/form/fieldset/div[1]/span/input'
password_input = '//*[@id="login_wrapper"]/form/fieldset/div[2]/span/input'
login_submit = '//*[@id="login_wrapper"]/form/fieldset/div[3]/div/span/input'
Falla_Temporal_Web = '//*[@id="login_wrapperMsg"]/ul/li/strong'

# Buscamos los campos, se les envia la informacion del login y se hace click
driver.find_element_by_xpath(username_input).send_keys("")
sleep(.2)
driver.find_element_by_xpath(password_input).send_keys("")
driver.find_element_by_xpath(login_submit).click()
sleep(1)

# Pantalla despues del login

# Definimos los elementos a buscar en esta pantalla

numero_de_credito_input = '//*[@id="numeroCredito"]'

boton_buscar_num_cred = '//*[@id="info"]/div/div/form/fieldset[1]/div/div/span[2]'

Monto_de_Constancia = '//*[@id="info"]/div/div/form/fieldset[2]/div/div[5]/span'
No_existe_credito_en_saldos = '//*[@id="info"]/div/div/form/fieldset[1]/div/div[2]/ul/li'


# Cargamos los datos a buscar
workbook = load_workbook(filedialog.askopenfilename())
worksheet = workbook.worksheets[0]
columna_a_importar = worksheet['C']
lista_de_creditos = [columna_a_importar[i].value for i in range(len(columna_a_importar))]
lista_de_creditos.pop(0)
contador = len(lista_de_creditos)
# convertimos el string a entero
# lista_de_creditos = [int(i) for i in lista_de_creditos]
print("Se encontraron " + str(len(lista_de_creditos)) + " numeros de credito")

inicio = time.time()
# Buscamos los datos
for i in lista_de_creditos:
    numero_credito = i
    # Escribimos el numero de credito en el input y lo guardamos en una lista, si el # de credito tiene menos de 10 digitos ingresamos un 0 antes
    if len(str(i)) < 10:
        driver.find_element_by_xpath(numero_de_credito_input).send_keys(0)
        driver.find_element_by_xpath(numero_de_credito_input).send_keys(numero_credito)
    else:
        driver.find_element_by_xpath(numero_de_credito_input).send_keys(numero_credito)

    Salida_keys.append(numero_credito)
    contador = contador - 1
    sleep(random.uniform(.1, .2))
    # hacemos click en buscar, obtenemos el monto del vale
    sleep(.1)
    driver.find_element_by_xpath(boton_buscar_num_cred).click()
    print("Faltan: " + str(contador) + " Folios")
    sleep(.1)
    try:
        monto_vale = driver.find_element_by_xpath(Monto_de_Constancia).text
        Salida_Values.append(monto_vale)
        print(str(i) + " " + str(monto_vale))
        sleep(.1)
        driver.find_element_by_xpath(numero_de_credito_input).clear()
        sleep(.1)
        driver.find_element_by_xpath(boton_buscar_num_cred).click()
        sleep(.1)

    except:
        monto_vale = driver.find_element_by_xpath(No_existe_credito_en_saldos).text
        sleep(.1)
        print(str(i) + " " + str(monto_vale))
        Salida_Values.append(monto_vale)
        driver.find_element_by_xpath(numero_de_credito_input).clear()
        sleep(.1)
# print(Salida_keys)
# print(Salida_Values)

Diccionario_salida = dict(zip(Salida_keys, Salida_Values))

# print(Diccionario_salida)

final = time.time()

elapsed = final - inicio

print("Proceso finalizado, por favor guarde el archivo con extension .XLSX")
print("Finalizado en : ", elapsed / 60, " Minutos.")
# Escribe los resultados en una hoja de excel
with xlsxwriter.Workbook(filedialog.asksaveasfilename()) as workbook:

    worksheet = workbook.add_worksheet()

    row = 0
    col = 0

    for key in Diccionario_salida.keys():
        worksheet.write(row, col, key)
        row += 1

    row = 0

    for item in Diccionario_salida.values():
        worksheet.write(row, col + 1, item)
        row += 1
# Cierra el explorador y finaliza tkinter

driver.close()
root.destroy()
